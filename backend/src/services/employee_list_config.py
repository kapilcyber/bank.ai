"""Employee list config and verification: app_config, company_employee_list table, CSV/Excel upload only."""
import csv
import io
import json
import os
from typing import Optional, Dict, Any, List
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, func, delete

from src.models.employee_list import AppConfig, CompanyEmployeeList
from src.utils.logger import get_logger

ALLOWED_EXTENSIONS = (".csv", ".xlsx", ".xls")

logger = get_logger(__name__)

CONFIG_KEY_ENABLED = "employee_verification_enabled"
CONFIG_KEY_LEFT_AFTER_UPLOAD = "employee_list_left_after_upload"


async def is_employee_list_available(db: AsyncSession) -> bool:
    """Return True if the uploaded list in the database has data for verification."""
    r_count = await db.execute(select(func.count(CompanyEmployeeList.id)))
    return (r_count.scalar() or 0) > 0


async def get_employee_list_config(db: AsyncSession) -> Dict[str, Any]:
    """Return enabled and count of employees in the uploaded list (stored in DB)."""
    q_enabled = select(AppConfig.value).where(AppConfig.key == CONFIG_KEY_ENABLED)
    r_enabled = await db.execute(q_enabled)
    enabled_val = (r_enabled.scalar_one_or_none() or "true").strip().lower()
    enabled = enabled_val in ("true", "1", "yes")

    r_count = await db.execute(select(func.count(CompanyEmployeeList.id)))
    count = r_count.scalar() or 0

    return {"enabled": enabled, "count": count}


async def set_employee_list_config(
    db: AsyncSession,
    enabled: Optional[bool] = None,
) -> None:
    """Update app_config for employee list (verification on/off). Only enabled is used."""
    if enabled is not None:
        await db.execute(
            delete(AppConfig).where(AppConfig.key == CONFIG_KEY_ENABLED)
        )
        db.add(AppConfig(key=CONFIG_KEY_ENABLED, value="true" if enabled else "false"))
    await db.commit()


async def set_left_employees_after_upload(db: AsyncSession, left_employees: List[Dict[str, Any]]) -> None:
    """Store the list of employees who were in the previous list but not in the new one (after an upload)."""
    await db.execute(delete(AppConfig).where(AppConfig.key == CONFIG_KEY_LEFT_AFTER_UPLOAD))
    db.add(AppConfig(key=CONFIG_KEY_LEFT_AFTER_UPLOAD, value=json.dumps(left_employees)))
    await db.commit()


async def get_left_employees_after_upload(db: AsyncSession) -> List[Dict[str, Any]]:
    """Return the stored list of employees who left (not in new list after last upload). Empty if none or no upload yet."""
    r = await db.execute(select(AppConfig.value).where(AppConfig.key == CONFIG_KEY_LEFT_AFTER_UPLOAD))
    val = r.scalar_one_or_none()
    if not val or not (val or "").strip():
        return []
    try:
        return json.loads(val)
    except (json.JSONDecodeError, TypeError):
        return []


def _normalize_header(name: str) -> str:
    """Normalize CSV/Excel header: strip, lowercase; map common names to employee_id, full_name, email."""
    s = (name or "").strip().lower().lstrip("\ufeff")
    if s == "full name":
        return "full_name"
    s = s.replace(" ", "_")
    if s in ("emp_no", "employee_no", "employee_number", "emp_id"):
        return "employee_id"
    if s in ("e-mail", "e_mail", "email_address"):
        return "email"
    if s == "name":
        return "full_name"
    return s


async def verify_employee_against_list(
    employee_id: str,
    email: str,
    db: AsyncSession,
) -> Optional[Dict[str, str]]:
    """
    Verify (employee_id, email) against the uploaded list in company_employee_list table.
    Returns {"full_name": "..."} if found, None otherwise.
    When verification is disabled, returns {"full_name": ""} to allow signup without list check.
    """
    config = await get_employee_list_config(db)
    enabled = config.get("enabled", True)

    if not enabled:
        return {"full_name": ""}

    eid = (employee_id or "").strip().upper()
    em = (email or "").strip().lower()
    if not eid or not em:
        return None
    q = select(CompanyEmployeeList).where(
        CompanyEmployeeList.employee_id == eid,
        CompanyEmployeeList.email == em,
    )
    result = await db.execute(q)
    row = result.scalar_one_or_none()
    if row:
        return {"full_name": (row.full_name or "").strip()}
    return None


def _parse_employee_file(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Parse CSV or Excel (employee_id, full_name, email). Returns list of dicts with employee_id, full_name, email.
    Raises ValueError on missing required columns or duplicate employee_id.
    """
    ext = (os.path.splitext(filename or "")[1] or "").lower()
    if ext not in ALLOWED_EXTENSIONS:
        raise ValueError(f"File must be CSV or Excel (.csv, .xlsx, .xls). Got: {ext or 'no extension'}")

    rows: List[Dict[str, Any]] = []
    seen_ids: set = set()
    required = {"employee_id", "email"}

    def validate_and_append(row_dict: Dict[str, str]) -> None:
        row_dict = {_normalize_header(k): (str(v).strip() if v is not None else "") for k, v in row_dict.items()}
        eid = (row_dict.get("employee_id") or "").strip()
        email = (row_dict.get("email") or "").strip()
        if not eid or not email:
            return
        if eid in seen_ids:
            raise ValueError(f"Duplicate employee_id: {eid}")
        seen_ids.add(eid)
        rows.append({
            "employee_id": eid.upper(),
            "full_name": (row_dict.get("full_name") or "").strip() or None,
            "email": email.lower(),
        })

    if ext == ".csv":
        decoded = file_content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(decoded))
        fieldnames = [_normalize_header(h) for h in (reader.fieldnames or [])]
        if not required.issubset(fieldnames):
            raise ValueError("File must have columns: employee_id, full_name (or full name), email")
        for row in reader:
            validate_and_append(row)
        return rows

    if ext == ".xlsx":
        from openpyxl import load_workbook
        # Use normal mode so we can read header then data reliably (read_only allows only one iterator)
        wb = load_workbook(io.BytesIO(file_content), read_only=False)
        ws = wb.active
        if ws is None:
            raise ValueError("Excel file has no sheet")
        # Get header from first row (row 1 in 1-based indexing)
        header_cells = list(ws.iter_rows(min_row=1, max_row=1))[0]
        fieldnames = [_normalize_header(str(c.value or "")) for c in header_cells]
        if not required.issubset(set(fieldnames)):
            raise ValueError(
                "File must have columns: employee_id, full_name (or full name), email. "
                f"Found: {', '.join(fieldnames) or 'none'}"
            )
        for row in ws.iter_rows(min_row=2):
            row_dict = {fieldnames[i]: row[i].value for i in range(min(len(fieldnames), len(row)))}
            validate_and_append(row_dict)
        return rows

    if ext == ".xls":
        import xlrd
        book = xlrd.open_workbook(file_contents=file_content)
        sheet = book.sheet_by_index(0)
        if sheet.nrows < 1:
            raise ValueError("Excel file has no header row")
        fieldnames = [_normalize_header(str(sheet.cell_value(0, c))) for c in range(sheet.ncols)]
        if not required.issubset(set(fieldnames)):
            raise ValueError("File must have columns: employee_id, full_name (or full name), email")
        for r in range(1, sheet.nrows):
            row_dict = {fieldnames[c]: sheet.cell_value(r, c) for c in range(sheet.ncols) if c < len(fieldnames)}
            validate_and_append(row_dict)
        return rows

    raise ValueError(f"Unsupported file type: {ext}")


async def replace_employee_list_from_csv(
    db: AsyncSession,
    file_content: bytes,
    filename: str = "upload.csv",
):
    """
    Parse CSV or Excel (employee_id, full_name, email), validate, overwrite company_employee_list table, bulk insert.
    Returns (count, removed): count of rows inserted, and list of {email, employee_id, full_name} for employees
    who were in the previous list but are not in the new list (they "left").
    """
    # Current list before replace (to compute who left)
    current_result = await db.execute(select(CompanyEmployeeList))
    current_rows = list(current_result.scalars().all())
    new_rows = _parse_employee_file(file_content, filename)
    new_emails = {(r["email"] or "").strip().lower() for r in new_rows}
    removed = [
        {"email": (r.email or "").strip().lower(), "employee_id": (r.employee_id or "").strip(), "full_name": (r.full_name or "").strip() or None}
        for r in current_rows
        if (r.email or "").strip().lower() and (r.email or "").strip().lower() not in new_emails
    ]

    await db.execute(delete(CompanyEmployeeList))
    for r in new_rows:
        db.add(CompanyEmployeeList(
            employee_id=r["employee_id"],
            full_name=r.get("full_name"),
            email=r["email"],
        ))
    await db.commit()
    return len(new_rows), removed
